import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = "/Users/sheikheddy/Projects/memesong/copy-strings.json"
OUT = "/Users/sheikheddy/Projects/memesong/ninefold-copy-sheet.xlsx"

rows = json.load(open(SRC))

wb = Workbook()

# ---------------------------------------------------------------- how to use
guide = wb.active
guide.title = "How to use"
ARIAL = "Arial"

guide["A1"] = "Ninefold — copy rewrite sheet"
guide["A1"].font = Font(name=ARIAL, size=16, bold=True)
guide["A2"] = "Every user-visible string in the game, pulled from index.html and src/game.js."
guide["A2"].font = Font(name=ARIAL, size=11)

notes = [
    ("", ""),
    ("Where to write", "Type your replacement in the yellow 'New text' column on the Strings tab. Leave a row blank to keep the current wording."),
    ("Two-line copy", "<br> marks a hard line break inside one string. Keep it where you want the line to split — most headings and couplets rely on it."),
    ("Values in braces", "${...} is filled in at runtime (a name, a number). Keep every placeholder that appears in the current text, or that value will vanish."),
    ("Newlines", "\\n in floating battle text is a line break, same idea as <br>."),
    ("Screen", "Which part of the game the string appears in, so you can rewrite one surface at a time."),
    ("Location", "file:line in the repo, if you want to see it in context."),
    ("Length", "Character count of the current text. Buttons and small labels get clipped, so stay near the current length for anything under ~30 characters."),
    ("", ""),
    ("Example", "Screen: Masthead · Current: THE SIGN WANTS A BODY<br>THE BODY WANTS A SIGN · New text: A NAME WANTS A THROAT<br>A THROAT WANTS A NAME"),
]
r = 4
for label, text in notes:
    if label:
        guide.cell(row=r, column=1, value=label).font = Font(name=ARIAL, size=11, bold=True)
        c = guide.cell(row=r, column=2, value=text)
        c.font = Font(name=ARIAL, size=11)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        guide.row_dimensions[r].height = 30
    r += 1

guide["A15"] = "Total strings"
guide["A15"].font = Font(name=ARIAL, size=11, bold=True)
guide["B15"] = len(rows)
guide["B15"].font = Font(name=ARIAL, size=11)

guide["A16"] = "Progress"
guide["A16"].font = Font(name=ARIAL, size=11, bold=True)
guide["B16"] = '=COUNTA(Strings!F2:F%d) & " of %d rewritten"' % (len(rows) + 1, len(rows))
guide["B16"].font = Font(name=ARIAL, size=11)
guide["C16"] = "(counts once your spreadsheet app opens the file)"
guide["C16"].font = Font(name=ARIAL, size=9, italic=True, color="7A7A7A")

guide.column_dimensions["A"].width = 18
guide.column_dimensions["B"].width = 96

# ------------------------------------------------------------------ strings
ws = wb.create_sheet("Strings")
headers = ["#", "Screen", "Element / field", "Location", "Current text", "New text", "Len", "Notes"]
head_fill = PatternFill("solid", fgColor="1D1821")
edit_fill = PatternFill("solid", fgColor="FFF6C2")
band_fill = PatternFill("solid", fgColor="F4F2EF")
thin = Side(style="thin", color="D9D5D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for i, h in enumerate(headers, start=1):
    c = ws.cell(row=1, column=i, value=h)
    c.font = Font(name=ARIAL, size=11, bold=True, color="FFFDF5")
    c.fill = head_fill
    c.alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 24

prev_screen = None
band = False
for idx, row in enumerate(rows, start=2):
    if row["screen"] != prev_screen:
        band = not band
        prev_screen = row["screen"]
    values = [
        idx - 1,
        row["screen"],
        row["element"],
        row["location"],
        row["text"],
        None,
        len(row["text"]),
        row.get("notes", ""),
    ]
    for col, v in enumerate(values, start=1):
        c = ws.cell(row=idx, column=col, value=v)
        c.font = Font(name=ARIAL, size=10)
        c.border = border
        c.alignment = Alignment(wrap_text=(col in (3, 5, 6, 8)), vertical="top")
        if col == 6:
            c.fill = edit_fill
        elif band:
            c.fill = band_fill
    ws.row_dimensions[idx].height = None

widths = {"A": 5, "B": 17, "C": 27, "D": 17, "E": 52, "F": 52, "G": 6, "H": 30}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
ws.freeze_panes = "E2"
ws.auto_filter.ref = "A1:H%d" % (len(rows) + 1)

wb.save(OUT)
print("wrote", OUT, "with", len(rows), "rows")
